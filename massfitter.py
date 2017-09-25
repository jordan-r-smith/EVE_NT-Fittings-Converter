"""This script is used to convert the EVE_NT ship fittings spreadsheet into an
easily importable XML file for use with the EVE Online client."""

from xml.etree import ElementTree as ET
from openpyxl import load_workbook

def indent(elem, level=0):
    """Pretty print the XML"""
    i = "\n" + level*"  "
    if len(elem):
        if not elem.text or not elem.text.strip():
            elem.text = i + "  "
        if not elem.tail or not elem.tail.strip():
            elem.tail = i
        for elem in elem:
            indent(elem, level + 1)
        if not elem.tail or not elem.tail.strip():
            elem.tail = i
    else:
        if level and (not elem.tail or not elem.tail.strip()):
            elem.tail = i

def build_fits():
    """Read in spreadsheet and generate XML"""

    # Load spreadsheet
    wb_fittings = load_workbook(filename='fittings.xlsx')
    xml_fittings = ET.Element("fittings")

    # Spreadsheet loop
    for sheet in wb_fittings.worksheets:
        if sheet.title != "Overall":
            # Columns loop
            for col in sheet.iter_cols(min_row=1, max_col=sheet.max_column, max_row=sheet.max_row):
                xml_single_fit = ET.SubElement(xml_fittings, "fitting")
                ET.SubElement(xml_single_fit, "description", value="")

                slots = ["low slot ", "med slot ", "hi slot ", "rig slot ", "cargo"]
                slots_index = 0
                slot_position = 0 # e.g. low slot 0, low slot 1, etc.

                drones = ["acolyte", "hornet", "hobgoblin", "warrior", "infiltrator", "vespa",
                          "hammerhead", "valkyrie", "praetor", "wasp", "ogre", "berserker",
                          "gecko", "bouncer", "curator", "garde", "warden", "maintenance bot"]

                # Rows loop
                for index, cell in enumerate(col):
                    # Grab fitting header
                    if index == 1:
                        fit_header = cell.value[1:-1].split(', ')
                        xml_single_fit.set("name", fit_header[1] + " " + fit_header[0])
                        ET.SubElement(xml_single_fit, "shipType", value=fit_header[0])
                    elif index > 1:
                        if cell.value:
                            # Parse drones and cargo items
                            if slots[slots_index] is "cargo":
                                slot_item = cell.value.lower()
                                if any(slot_item.find(d) >= 0 for d in drones):
                                    cargo_item = cell.value.split(' x')
                                    ET.SubElement(xml_single_fit, "hardware", qty=cargo_item[1],
                                                  slot="drone bay", type=cargo_item[0])
                                    slot_position += 1
                                else:
                                    cargo_item = cell.value.split(' x')
                                    ET.SubElement(xml_single_fit, "hardware", qty=cargo_item[1],
                                                  slot=slots[slots_index], type=cargo_item[0])
                                    slot_position += 1
                            # Parse everything else
                            else:
                                slot_item = cell.value.split(', ')[0]

                                if "[empty " not in slot_item:
                                    ET.SubElement(xml_single_fit, "hardware",
                                                  slot=slots[slots_index] + str(slot_position),
                                                  type=slot_item.lstrip())

                                slot_position += 1
                        else:
                            slots_index += 1
                            slot_position = 0

    # Beautify and write to XML
    indent(xml_fittings)
    tree = ET.ElementTree(xml_fittings)
    tree.write("EVE_NT_Cup_Fittings.xml", xml_declaration=True, encoding='utf-8', method="xml")
    print "Completed"

if __name__ == "__main__":
    build_fits()
